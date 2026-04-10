import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\91992\Downloads\Leave Summary Report (3).xlsx', data_only=True)
ws = wb.active

# Col: 1=EmpNo, 2=Name, 8=Open COF, 10=Open PL
# 13=COF Grant, 14=PL Grant
# 16=COF Availed, 17=LOP Availed, 18=PL Availed, 20=LOP (balance col)

# Employees already imported (from XLS day-wise)
imported = {
    'COLOR034','COLOR026','COLOR064','COLOR120','COLOR146','COLOR057',
    'COLOR013','COLOR022','COLOR170','COLOR171','COLOR176','COLOR191',
    'COLOR117','COLOR188','COLOR190','COLOR184','COLOR167','COLOR182',
    'COLOR185','COLOR194','COLOR193','COLOR115','COLOR195','COLOR047'
}

records = []
missing = []

for r in range(8, ws.max_row + 1):
    emp_no = str(ws.cell(r, 1).value or '').strip()
    if not emp_no.startswith('COLOR'):
        continue
    name     = str(ws.cell(r, 2).value or '').strip()
    cof_used = float(ws.cell(r, 16).value or 0)
    lop_used = float(ws.cell(r, 17).value or 0)
    pl_used  = float(ws.cell(r, 18).value or 0)

    total_used = cof_used + lop_used + pl_used
    records.append({'empNo': emp_no, 'name': name, 'cofUsed': cof_used, 'lopUsed': lop_used, 'plUsed': pl_used})

    if emp_no not in imported and total_used > 0:
        missing.append({'empNo': emp_no, 'name': name, 'cofUsed': cof_used, 'lopUsed': lop_used, 'plUsed': pl_used})
        print(f"  MISSING {emp_no:10s} | {name:35s} | PL={pl_used} COF={cof_used} LOP={lop_used}")

print(f"\nTotal employees: {len(records)}")
print(f"Already imported: {len(imported)}")
print(f"Missing (have leaves but not imported): {len(missing)}")

with open(r'D:\Activity Report Software\server\missing_leaves.json', 'w') as f:
    json.dump(missing, f, indent=2)
# Also save full used data for balance update
with open(r'D:\Activity Report Software\server\all_used.json', 'w') as f:
    json.dump(records, f, indent=2)
print("Saved missing_leaves.json and all_used.json")
