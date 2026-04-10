import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\91992\Downloads\Leave Balance As On A Day (3).xlsx', data_only=True)
ws = wb.active

print("Col layout: SL | EmpNo | Name | MgrNo | MgrName | COF | LOP | PL | RestHol | CF+AprPL")
print("\n--- All employee rows ---")
MAX_CF = 6

employees = []
for r in range(5, ws.max_row + 1):
    emp_no = str(ws.cell(r, 2).value or '').strip()
    if not emp_no or not emp_no.startswith('COLOR'):
        continue
    name     = str(ws.cell(r, 3).value or '').strip()
    cof_raw  = ws.cell(r, 6).value  # Comp-Off balance
    lop_raw  = ws.cell(r, 7).value  # Loss of Pay
    pl_raw   = ws.cell(r, 8).value  # Privilege Leave
    j_raw    = ws.cell(r, 10).value # CF + Apr PL

    cof_bal  = float(cof_raw) if cof_raw else 0.0
    j_val    = float(j_raw)   if j_raw   else 0.0

    # J = carry_forward_from_2024 + 1 (April PL bucket)
    pl_carry = round(j_val - 1, 2)          # remove April bucket
    cf_open  = round(min(pl_carry, MAX_CF), 2)
    lapsed   = round(max(pl_carry - MAX_CF, 0), 2)

    employees.append({
        'empNo': emp_no, 'name': name,
        'cofOpen': max(cof_bal, 0),  # COF carry-forward (only positive)
        'cfOpen': cf_open,           # PL carry → CF.opening (capped 6)
        'lapsed': lapsed,
    })
    flag = ' *** LAPSED' if lapsed > 0 else ''
    print(f"  {emp_no:10s} | {name:35s} | COF carry={max(cof_bal,0):4.1f} | CF.open={cf_open:4.1f} | Lapsed={lapsed:4.1f}{flag}")

print(f"\nTotal: {len(employees)} employees")
lapsed_emps = [e for e in employees if e['lapsed'] > 0]
print(f"Employees with lapsed PL: {len(lapsed_emps)}")

import json
with open(r'D:\Activity Report Software\server\opening_balances.json', 'w') as f:
    json.dump(employees, f, indent=2)
print("Saved to opening_balances.json")
