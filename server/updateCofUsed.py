import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\91992\Downloads\Leave Summary Report (3).xlsx', data_only=True)
ws = wb.active

# Col 1=EmpNo, 13=COF Grant, 16=COF Availed
records = []
for r in range(8, ws.max_row + 1):
    emp_no = str(ws.cell(r, 1).value or '').strip()
    if not emp_no.startswith('COLOR'):
        continue
    name      = str(ws.cell(r, 2).value or '').strip()
    cof_grant = ws.cell(r, 13).value
    cof_used  = ws.cell(r, 16).value

    cof_grant = float(cof_grant) if cof_grant and str(cof_grant).strip() not in ('', 'None') else 0.0
    cof_used  = float(cof_used)  if cof_used  and str(cof_used).strip()  not in ('', 'None') else 0.0

    records.append({'empNo': emp_no, 'name': name, 'cofGrant': cof_grant, 'cofUsed': cof_used})
    if cof_used > 0:
        print(f"  {emp_no:10s} | {name:35s} | grant={cof_grant:5.1f} | used={cof_used:5.1f}")

print(f"\nTotal: {len(records)}, with COF used > 0: {sum(1 for r in records if r['cofUsed'] > 0)}")
with open(r'D:\Activity Report Software\server\cof_used.json', 'w') as f:
    json.dump(records, f, indent=2)
print("Saved cof_used.json")
