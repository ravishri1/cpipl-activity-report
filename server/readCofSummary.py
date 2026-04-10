import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\91992\Downloads\Leave Summary Report (3).xlsx', data_only=True)
ws = wb.active

print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
print("\n--- First 8 rows (headers) ---")
for r in range(1, 9):
    row = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column+1)]
    print(f"Row {r}: {row}")

print("\n--- First 15 data rows ---")
for r in range(9, 24):
    row = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column+1)]
    print(f"Row {r}: {row}")
