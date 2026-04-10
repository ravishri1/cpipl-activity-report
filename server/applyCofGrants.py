import openpyxl, json

wb = openpyxl.load_workbook(r'C:\Users\91992\Downloads\Leave Summary Report (3).xlsx', data_only=True)
ws = wb.active

# Col layout (1-indexed):
# 1=EmpNo, 2=Name, 8=Opening COF, 9=Opening LOP, 10=Opening PL
# 13=COF Eligibility (granted), 14=PL Eligibility
# 16=COF Availed, 17=LOP Availed, 18=PL Availed

records = []
for r in range(8, ws.max_row + 1):
    emp_no = str(ws.cell(r, 1).value or '').strip()
    if not emp_no.startswith('COLOR'):
        continue
    name       = str(ws.cell(r, 2).value or '').strip()
    cof_grant  = ws.cell(r, 13).value  # COF Eligibility = actual grants this FY
    pl_elig    = ws.cell(r, 14).value  # PL Eligibility

    cof_grant = float(cof_grant) if cof_grant and str(cof_grant).strip() not in ('', 'None') else 0.0

    records.append({'empNo': emp_no, 'name': name, 'cofGrant': cof_grant})
    if cof_grant > 0:
        print(f"  {emp_no:10s} | {name:35s} | COF Grant = {cof_grant}")

print(f"\nTotal: {len(records)} employees, {sum(1 for r in records if r['cofGrant'] > 0)} with COF grants")

with open(r'D:\Activity Report Software\server\cof_grants.json', 'w') as f:
    json.dump(records, f, indent=2)
print("Saved to cof_grants.json")
