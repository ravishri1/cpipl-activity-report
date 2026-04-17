import openpyxl
import json
from datetime import datetime

MUSTER_PATH = r'C:\Users\91992\Downloads\Attendance Muster Report (3).xlsx'

wb = openpyxl.load_workbook(MUSTER_PATH, data_only=True)
ws = wb.active
rows = []
cur_emp = None

for r in range(1, ws.max_row + 1):
    v = [str(ws.cell(r, c).value or '').strip().strip("'") for c in range(1, 19)]
    v0 = v[0]
    if v0.startswith('COLOR') and len(v0) == 8:
        cur_emp = v0
    elif cur_emp and v[0].replace('.0', '').isdigit():
        date_val = ws.cell(r, 2).value
        if isinstance(date_val, datetime):
            rows.append({
                'emp': cur_emp,
                'date': date_val.strftime('%Y-%m-%d'),
                's1': v[2].upper(),
                's2': v[3].upper()
            })

print(json.dumps(rows))
