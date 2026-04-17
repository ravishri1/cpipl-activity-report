"""
Parse June and July 2025 Excel salary registers and output as JSON
for comparison with EOD-generated payslips.
"""
import openpyxl
import json

FILES = {
    '2025-06': r'C:\Users\91992\Downloads\colorpapersindiaprivatelimitedsalaryregisterfy20252\Final Salary Register June 25 -Final.xlsx',
    '2025-07': r'C:\Users\91992\Downloads\colorpapersindiaprivatelimitedsalaryregisterfy20252\Salary for the month of Jul - 2025 - Software.xlsx',
}

# Column indices (0-based) for each month
JUNE_COLS = {
    'empId': 0, 'name': 1, 'lop': 6, 'effDays': 7,
    'basic': 8, 'gross': 15, 'pf': 16, 'esi': 17, 'pt': 18,
    'loan': 20, 'advance': 21, 'lwf': 22, 'totalDed': 23, 'netPay': 24
}
JULY_COLS = {
    'empId': 0, 'name': 1, 'lop': 7, 'effDays': 8,
    'basic': 9, 'gross': 15, 'pf': 16, 'esi': 17, 'pt': 18,
    'loan': 19, 'totalDed': 20, 'netPay': 21
}

def parse_num(v):
    try:
        return float(str(v).replace(',', '').strip()) if v else 0
    except:
        return 0

def parse_month(filepath, cols, data_start_row=4):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []
    for r in range(data_start_row, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 27)]
        empId = str(vals[cols['empId']] or '').strip()
        if not empId.startswith('COLOR'):
            continue
        rows.append({
            'empId': empId,
            'name': str(vals[cols['name']] or '').strip(),
            'lop': parse_num(vals[cols['lop']]),
            'effDays': parse_num(vals[cols['effDays']]),
            'basic': parse_num(vals[cols['basic']]),
            'gross': parse_num(vals[cols['gross']]),
            'pf': parse_num(vals[cols['pf']]),
            'esi': parse_num(vals[cols.get('esi', -1)] if cols.get('esi', -1) >= 0 else 0),
            'pt': parse_num(vals[cols['pt']]),
            'loan': parse_num(vals[cols.get('loan', -1)] if cols.get('loan', -1) >= 0 else 0),
            'advance': parse_num(vals[cols.get('advance', -1)] if cols.get('advance', -1) >= 0 else 0),
            'lwf': parse_num(vals[cols.get('lwf', -1)] if cols.get('lwf', -1) >= 0 else 0),
            'totalDed': abs(parse_num(vals[cols['totalDed']])),
            'netPay': parse_num(vals[cols['netPay']]),
        })
    return rows

result = {}
for month, filepath in FILES.items():
    cols = JUNE_COLS if month == '2025-06' else JULY_COLS
    result[month] = parse_month(filepath, cols)

print(json.dumps(result))
